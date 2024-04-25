#!/usr/bin/env python
# cbaero_surrogate.py - Tyler Adams 1/12/2023
"""
Builds Kriging surrogate model to predict aerodynamic coefficients for given geometry based on angle of attack,
mach number, & dynamic pressure. Employs latin hypercube sampling to generate number of training & testing
points for model. Runs CBAero for each of those points, saves output data, & reads in that data to build
Kriging model. Generates data text files and plots showing model error.

"""

import os
import sys
import subprocess
import argparse
import shutil
import pickle
import logging
import glob

import numpy as np
import matplotlib.pyplot as plt
from dataclasses import dataclass, field

# Surrogate modeling
from smt.surrogate_models import KRG
from smt.applications.mfk import MFK
from sklearn.metrics import mean_squared_error
from pyDOE2 import lhs

from openpyxl import load_workbook

import random

OTIS = '.otis'
TECPLOT = '_tecplot.dat'

BASE_OTIS_DIR = './otis'
CBAERO_OUTPUT = 'cbaero_output.txt'
PTS_CL_CD_FILE = 'pts_cl_cd.pkl'
TEST_TO_TRAIN = 0.33
RAND_STATE = 25  # Latin hypercube random seed

# Correlation names
CORR_LIFT = 'Lift'
CORR_DRAG = 'Drag'

# Correlation counts
CORR_COUNT = {CORR_LIFT: 0.001, CORR_DRAG: 0.0001}

PSF_TO_BARS = 0.000479

CBAERO_MACH = '1'
CBAERO_DYN_PRESS = '1'
CBAERO_ANGLE_ATTACK = '1'

# CBAero run modes
TRAIN = '_train_'
TEST = '_test_'

# CbaeroSurrogate Run modes
SERIAL = 'serial'
PARALLEL_START = 'pstart'
PARALLEL_END = 'pend'

# Fit krig
CL_MODE = "cl"
CD_MODE = "cd"
QDOT_MODE = "qdot"

# Model
TAG_REGIONS = 'TaggedRegions'
MESH_EXT = '.msh'
SETUP_EXT = '.stp'

# Correlation Function
CORR_FUNC = "squar_exp"

# Q dot options
QDOT_BOTH = 0  # cb2otis & adb2tecplot run, either all cl/cd/qdot from user's' cokrig file -or- not running cokriging
QDOT_CB2OTIS = 1  # cb2otis runs, (not adb2tecplot), only cl/cd from cokrig file, user has no qdot data
QDOT_ADB2TECPLOT = 2  # adb2tecplot runs (not cb2otis), qdot surrogate built w/kriging, cl/cd models use cokriging w/resp. train/test sets


@dataclass
class TrainTestValues:
    """Store test and train values."""
    train: np.ndarray = field(default=None)
    test: np.ndarray = field(default=None)


@dataclass
class HighFidelityData:
    """Store high fidelity data read from cokrig file."""
    x: TrainTestValues
    cl: TrainTestValues
    cd: TrainTestValues


@dataclass
class HighFidelityTrain:
    """Store training values selected from HF data."""
    x: np.ndarray
    cl: np.ndarray
    cd: np.ndarray


class CbaeroSurrogate:
    """Build latin hypercube, run CBAero, create Krigging model, draw plots"""

    def __init__(self, model_path, min_alpha, max_alpha, min_mach, max_mach, min_q, max_q, cokrig_file, save_as):

        # Model name, files,& dirs
        self.model_dir, self.model_filename = os.path.split(model_path)
        self.model_name, _ = os.path.splitext(self.model_filename)

        # Parameters
        self.min_alpha = min_alpha
        self.max_alpha = max_alpha
        self.min_mach = min_mach
        self.max_mach = max_mach
        self.min_q = min_q
        self.max_q = max_q

        # Cokrig file
        self.cokrig_file = cokrig_file

        # Save file as
        self.save_as = save_as

        # Dir in which to place .otis files - NOTE: parallel runs will create subdirs in this
        self.base_otis_dir = BASE_OTIS_DIR
        os.makedirs(self.base_otis_dir, exist_ok=True)  # Create dir if needed
        self.otis_dir = self.base_otis_dir  # Set current otis dir

        # CBAero output file handle
        self.cbaero_output = open(CBAERO_OUTPUT, 'a')

        # Maintain lists for num training pts, lift & drag counts - used in plotting
        self.clear_lists()

    def __del__(self):
        self.cbaero_output.close()

    def change_otis_dir(self, otis_dir):
        """Change the dir for storing .otis files."""
        self.otis_dir = otis_dir
        os.makedirs(self.otis_dir, exist_ok=True)  # Create dir if needed

    def get_hf_data(self):
        """Read high fidelity data from cokrig file."""

        # Step 1: Extract data from worksheet

        # NOTE Assumes FOUR columns: alpha, mach, cl, cd
        columns = range(1,5,1) # 1,2,3,4
        ws = load_workbook(self.cokrig_file).active
        alphas = []
        machs = []
        cl = []
        cd = []

        for col in ws.iter_cols(min_row=1, min_col=columns[0], max_col=columns[0], values_only=True):
            for cell in col:
                alphas.append(cell)

        for col in ws.iter_cols(min_row=1, min_col=columns[1], max_col=columns[1], values_only=True):
            for cell in col:
                machs.append(cell)

        for col in ws.iter_cols(min_row=1, min_col=columns[2], max_col=columns[2], values_only=True):
            for cell in col:
                cl.append(cell)

        for col in ws.iter_cols(min_row=1, min_col=columns[3], max_col=columns[3], values_only=True):
            for cell in col:
                cd.append(cell)

        # Step 2: Parse data, load into output data structure

        # Temp vars to aid extraction
        x_hf = np.array(list(zip(alphas,machs)))
        cd_hf = np.array(cd)
        cl_hf = np.array(cl)

        hf_data = HighFidelityData(TrainTestValues(), TrainTestValues(), TrainTestValues())  # Return structure

        hf_data.x.test = x_hf[1::2]  # select every other element starting at 1
        hf_data.x.train = x_hf[::2]  # select every other element starting at 0

        hf_data.cd.test = cd_hf[1::2] # select every other element starting at 1
        hf_data.cd.train = cd_hf[::2] # select every other element starting at 0

        hf_data.cl.test = cl_hf[1::2] # select every other element starting at 1
        hf_data.cl.train = cl_hf[::2] # select every other element starting at 0

        hf_data.cd.test = hf_data.cd.test.reshape(-1,1)
        hf_data.cd.train = hf_data.cd.train.reshape(-1,1)
        hf_data.cl.test = hf_data.cl.test.reshape(-1,1)
        hf_data.cl.train = hf_data.cl.train.reshape(-1,1)

        return hf_data

    def latin_hypercube(self, num_training_pts, rand_state):
        """Get LH sampling for data"""
        lb = np.array([self.min_alpha, self.min_mach])
        ub = np.array([self.max_alpha, self.max_mach])
        num_testing_pts = num_training_pts * TEST_TO_TRAIN
        x_train_lhs = lhs(n=2, samples=num_training_pts, criterion="cm", random_state=rand_state)
        x_train_lhs = lb + (ub - lb)*x_train_lhs
        x_test_lhs = lhs(n=2, samples=int(num_testing_pts), criterion="cm", random_state=rand_state)
        x_test_lhs = lb + (ub - lb)*x_test_lhs
        return x_train_lhs, x_test_lhs

    def latin_hypercube_qdot(self, num_training_pts, rand_state):
        """Get LH sampling for data"""
        lb = np.array([self.min_alpha, self.min_mach, self.min_q])
        ub = np.array([self.max_alpha, self.max_mach, self.max_q])
        num_testing_pts = num_training_pts * TEST_TO_TRAIN
        x_train_lhs = lhs(n=3, samples=num_training_pts, criterion="cm", random_state=rand_state)
        x_train_lhs = lb + (ub - lb)*x_train_lhs
        x_test_lhs = lhs(n=3, samples=int(num_testing_pts), criterion="cm", random_state=rand_state)
        x_test_lhs = lb + (ub - lb)*x_test_lhs
        return x_train_lhs, x_test_lhs

    def get_latin_hypercube_cokriging(self, num_training, rand_state):
        """Get the LH Sampling for the data"""
        lb = np.array([self.min_alpha, self.min_mach])
        ub = np.array([self.max_alpha, self.max_mach])
        x_lf_lhs = lhs(n=2, samples=num_training, criterion="cm", random_state=rand_state)
        x_lf_lhs = lb + (ub - lb)*x_lf_lhs
        return x_lf_lhs

    def run_cbaero(self, mode, iterations, lhs_matrix, qdot_opt, task_id):
        """Execute CBAero binaries using modified input file"""

        # Set omp option for CBAero

        omp_threads = os.environ.get('OMP_NUM_THREADS')
        mkl_threads = os.environ.get('MKL_NUM_THREADS')

        if omp_threads is not None:
            threads_list = ['-omp', omp_threads]
        elif mkl_threads is not None:
            threads_list = ['-omp', mkl_threads]
        else:
            threads_list = []

        # Temporarily switch to otis dir so cbaero will put output files there
        curr_dir = os.getcwd()
        os.chdir(self.otis_dir)

        for i in range(iterations):
            logging.info(f'Mode: {mode}, Run: {i+1}')
            self.cbaero_output.write(f'\n\nTask: {task_id}, Mode: {mode}, Run: {i+1}\n\n')
            self.cbaero_output.flush()

            # Update input file

            with open(self.model_filename, "r") as input_file:
                lines = input_file.readlines()

            if lhs_matrix[i][1] < 1.2:
                lines[8] = 'flotyp =       ' + str(3) + '\n'
            else:
                lines[8] = 'flotyp =       ' + str(0) + '\n'

            lines[14] = CBAERO_MACH + '\n'
            lines[15] = ' ' + str(lhs_matrix[i][1]) + '\n'

            lines[17] = CBAERO_DYN_PRESS + '\n'

            if qdot_opt == QDOT_CB2OTIS:
                lines[18] = ' ' + str(random.uniform(self.min_q,self.max_q)) + '\n'
            else:
                lines[18] = ' ' + str(lhs_matrix[i][2]) + '\n'

            lines[20] = CBAERO_ANGLE_ATTACK + '\n'
            lines[21] = ' ' + str(lhs_matrix[i][0]) + '\n'

            with open(self.model_filename, "w") as input_file:
                input_file.writelines(lines)

            # Exec cbaero, cb2otis, and adb2tecplot - NOTE Assumes CBAero bins in user's PATH
            for step in range(1,4):

                if step == 1:
                    cmd = ['cbaero'] + threads_list + [self.model_name]
                elif step == 2 and qdot_opt != QDOT_ADB2TECPLOT:  # step 2
                    cmd = ['cb2otis', self.model_name]
                elif step == 3 and qdot_opt != QDOT_CB2OTIS:  # step 3
                    cmd = ['adb2tecplot', self.model_name, '1', '1', '1']

                result = subprocess.run(cmd, stdout=self.cbaero_output, text=True)

                # Make sure run was successful
                if result.returncode != 0:
                    logging.critical(f"{cmd} failed with return code: {result.returncode}")
                    sys.exit(1)

            if qdot_opt != QDOT_ADB2TECPLOT:
                # Rename otis file to save it
                os.rename(self.model_name+OTIS, self.model_name+mode+str(i)+OTIS)

            if qdot_opt != QDOT_CB2OTIS:
                # Rename tecplot file to make it easier to read
                tecplot_files = glob.glob(f'{self.model_name}.%*')
                tecplot_filename = tecplot_files[0]
                os.rename(tecplot_filename, self.model_name+mode+str(i)+TECPLOT)

        os.chdir(curr_dir)  # Return to orig dir

    def read_otis(self, filename):
        """Extract data from .otis file."""
        NUM_ALPHAS = 1  # count of angles of attack being tested
        NUM_MACH = 1  # count of mach numbers being tested
        NUM_Q = 1  # count of dynamic pressures being tested

        alpha_list = []  # Angles of attack being tested
        mach_list = []  # Mach numbers being tested
        cl_array = np.zeros((NUM_MACH * NUM_Q, NUM_ALPHAS))  # rows: amt of mach #s * amt of qs, cols: num alphas
        cd_array = np.zeros((NUM_MACH * NUM_Q, NUM_ALPHAS))  # rows = amt of mach #s * amt of qs, cols = num alphas

        # Read in each line of .otis file to extract data

        logging.debug(f'Reading otis file {os.path.join(self.otis_dir, filename)}')

        with open(os.path.join(self.otis_dir, filename)) as file:

            # Copy values from selected lines into corresponding arrays
            for i, text in enumerate(file):
                terms = text.split()  # "term term   term..." separated by any whitespace

                # Consider each "word" on current line
                for j in range(len(terms)):

                    if i == 10:  # Line 11: list of alphas
                        alpha_list.append(float(terms[j]))
                    elif i == 14:  # Line 15: list of mach #s
                        mach_list.append(float(terms[j]))
                    elif 20 <= i <= (20 + NUM_MACH*NUM_Q - 1):
                        # Line 21: begin array of Cls, num rows for this array is num_mach*num_q
                        cl_array[i-20][j] = terms[j]  # num_alpha*num_mach*num_q values
                    elif ((20 + NUM_MACH*NUM_Q - 1) + 22) <= i <= (((20 + NUM_MACH*NUM_Q - 1) + 22) + NUM_MACH*NUM_Q - 1):
                        # Lines 20 + num_mach*num_q -1 line begins array of Cds, num rows is same as before
                        cd_array[i-((20 + NUM_MACH*NUM_Q - 1) + 22)][j] = terms[j]  # num_alpha*num_mach*num_q values

        return alpha_list, mach_list, cl_array, cd_array

    def tecplot_reader(self, file):
        """Tecplot reader."""
        arrays = []
        Qdot_total=[]

        with open(os.path.join(self.otis_dir, file)) as a:

            for idx, line in enumerate(a.readlines()):
                if idx < 3:
                    continue
                elif idx >= 3:
                    arrays.append([float(s) for s in line.split()])

        for i in range(len(arrays)):

            if len(arrays[i])>3:
                Qdot_total.append(arrays[i][15])

        Qdot_max = max(Qdot_total) / 1000000    #now in W/mm^2
        os.remove(os.path.join(self.otis_dir, file))
        return Qdot_max

    def extract_otis(self, num_training_pts, qdot_opt):
        """Convert data from .otis to 2D arrays for kriging."""
        x = TrainTestValues(train=np.zeros((num_training_pts, 2)), test=np.zeros((int(num_training_pts*TEST_TO_TRAIN), 2)))
        cl = TrainTestValues(train=np.zeros((num_training_pts, 1)), test=np.zeros((int(num_training_pts*TEST_TO_TRAIN), 1)))
        cd = TrainTestValues(train=np.zeros((num_training_pts, 1)), test=np.zeros((int(num_training_pts*TEST_TO_TRAIN), 1)))
        qdot = TrainTestValues(train=np.zeros((num_training_pts, 1)), test=np.zeros((int(num_training_pts*TEST_TO_TRAIN), 1)))

        for i in range(num_training_pts):

            if qdot_opt != QDOT_CB2OTIS:
                alphas, machs, cl_temp, cd_temp = self.read_otis(self.model_name+TRAIN+str(i)+OTIS)
                x.train[i][0] = alphas[0]
                x.train[i][1] = machs[0]

                cl.train[i][0] = cl_temp[0][0]
                cd.train[i][0] = cd_temp[0][0]

            qdot.train[i][0] = self.tecplot_reader(self.model_name+TRAIN+str(i)+TECPLOT)

        for i in range(int(num_training_pts * TEST_TO_TRAIN)):

            if qdot_opt != QDOT_CB2OTIS:
                alphas, machs, cl_temp, cd_temp = self.read_otis(self.model_name+TEST+str(i)+OTIS)
                x.test[i][0] = alphas[0]
                x.test[i][1] = machs[0]

                cl.test[i][0] = cl_temp[0][0]
                cd.test[i][0] = cd_temp[0][0]

            qdot.test[i][0] = self.tecplot_reader(self.model_name+TEST+str(i)+TECPLOT)

        return x, cl, cd, qdot

    def extract_otis_cokrig(self, num_training_pts):
        """Convert data from .otis to 2D arrays for kriging."""
        x_train = np.zeros((num_training_pts, 2))
        cl_train = np.zeros((num_training_pts, 1))
        cd_train = np.zeros((num_training_pts, 1))

        for i in range(num_training_pts):
            alphas, machs, cl_temp, cd_temp = self.read_otis(self.model_name+'_train_'+str(i)+OTIS)

            x_train[i][0] = alphas[0]
            x_train[i][1] = machs[0]

            cl_train[i][0] = cl_temp[0][0]
            cd_train[i][0] = cd_temp[0][0]

        return x_train, cl_train, cd_train

    #### choose_correlation is no longer used, instead "squar_exp" will always be chosen because its a smooth function. ####
    #### However, this function will remain here in case we want to reinstate it at a later date. 03/11/2024 - T. Adams ####
    @staticmethod
    def choose_correlation(num_training_pts, x_train, x_test, coe_train, coe_test, name, count_type):
        """Choose best correlation function for kriging."""

        # Test diff correlation functions to see which is best for this set of data

        correlation_funcs = ["squar_exp", "matern52", "matern32", "abs_exp"]
        rmse_list = []

        for func in correlation_funcs:

            # Fit kriging
            sm = KRG(theta0=np.array([1e-2, 1e-2]), corr=func, theta_bounds=[1e-6, 1e2], print_global=False)
            sm.set_training_values(x_train, coe_train)
            sm.train()
            logging.debug(f"Choose correlation: optimal theta found : {sm.optimal_theta}")  # theta value found

            test_pred = sm.predict_values(x_test) # Predict at test values
            sm.predict_variances(x_test)  # Predict uncertainty at test values

            # Calc RMSE
            rmse = np.sqrt(mean_squared_error(coe_test, test_pred))
            rmse_list.append(rmse)
            logging.debug(f'RMSE = {rmse_list}')

        best_rmse_index = rmse_list.index(min(rmse_list))  # Find best (lowest) RMSE value
        best_corr = correlation_funcs[best_rmse_index]  # Use index of best RMSE value to get right function

        # Create plots, save to files

        if name == "Lift" or name == "Drag":
            _, ax = plt.subplots()
            ax.bar(correlation_funcs, rmse_list, width=0.6)
            ax.set_xlabel("Kriging Correlation Functions", fontsize=14)
            ax.set_ylabel("RMSE", fontsize=14)
            ax.set_title("RMSE of Kriging Correlation Functions (" + name + ")", fontsize=15)

            plt.savefig(f'RMSE_{name}_{num_training_pts}pts')
            plt.close()

            _, ax = plt.subplots()
            ax.bar(correlation_funcs, np.divide(rmse_list, count_type), width=0.6, color='orange')
            ax.set_xlabel("Kriging Correlation Functions", fontsize=14)
            ax.set_ylabel("RMSE in " + name + " Counts", fontsize=14)
            ax.set_title("RMSE of Kriging Correlation Functions in " + name + " Counts", fontsize=15)

            plt.savefig(f'RMSE_{name}_Counts_{num_training_pts}pts')
            plt.close()

        else:
            _, ax = plt.subplots()
            ax.bar(correlation_funcs, rmse_list, width=0.6)
            ax.set_xlabel("Kriging Correlation Functions", fontsize=14)
            ax.set_ylabel("NRMSE", fontsize=14)
            ax.set_title("NRMSE of Kriging Correlation Functions (" + name + ")", fontsize=15)

            plt.savefig(f'NRMSE_{name}_{num_training_pts}pts')
            plt.close()

        return best_corr

    def fit_krig(self, mode, x, other, best_corr, num_training_pts, qdot_opt):  # TODO Better name for "other"?
        """Fit kriging for Cl, Cd, or q."""
        if qdot_opt == QDOT_CB2OTIS:   #qdot
            sm = KRG(theta0=np.array([1e-2, 1e-2, 1e-2]), corr=best_corr, theta_bounds=[1e-6, 1e2],
                                    print_global=False)
        else:   #Cl or Cd
            sm = KRG(theta0=np.array([1e-2, 1e-2]), corr=best_corr, theta_bounds=[1e-6, 1e2],
                                    print_global=False)

        sm.set_training_values(x.train, other.train)
        sm.train()

        logging.debug(f"Fit krig {mode}: Optimal theta found : {sm.optimal_theta}")

        # Predict at test values
        test_pred = sm.predict_values(x.test)

        # Log predicted uncertainty at test values
        logging.info(f"Fit krig {mode}: Average Uncertainty = {np.mean(sm.predict_variances(x.test))}")

        # Calc root-mean-square error
        rmse = np.sqrt(mean_squared_error(other.test, test_pred))
        logging.info(f"Fit krig {mode}: RMSE = {rmse}")

        # Calc count

        if mode == CL_MODE:
            count = rmse / CORR_COUNT[CORR_LIFT]
        elif mode == CD_MODE:
            count = rmse / CORR_COUNT[CORR_DRAG]
        else: #Qdot
            count = rmse / np.ptp(other.test)

        logging.info(f"Fit krig {mode}: Count = {count}")

        if self.save_as is not None:

            with open(self.save_as + "_" + mode + "_" + str(num_training_pts) + "pts.pkl", "wb") as f:
                pickle.dump(sm, f)

        return sm, count

    def fit_cokrig(self, mode, x_lf, x_hf, x_test, coe_lf_train, coe_hf_train, coe_test, num_training_pts):

        sm = MFK(theta0=x_lf.shape[1] * [1.0], print_global=False)
        sm.set_training_values(x_lf, coe_lf_train, name=0)
        sm.set_training_values(x_hf, coe_hf_train)
        sm.train()

        # Predict at test values
        test_pred = sm.predict_values(x_test)

        # Calculating rmse
        rmse = np.sqrt(mean_squared_error(coe_test, test_pred))

        if mode == CL_MODE:
            count = rmse / CORR_COUNT[CORR_LIFT]
        else:
            count = rmse / CORR_COUNT[CORR_DRAG]

        #Save the Cokriging Model
        if self.save_as is not None:

            with open(self.save_as + "_" + mode + "_" + str(num_training_pts) + "pts.pkl", "wb") as f:
                pickle.dump(sm, f)

        return sm, count

    def plot_convergence(self):
        _, ax = plt.subplots()
        ax.plot(self.train_pts_list, self.cl_count_list, marker="o")
        ax.set_xlabel("Training Points", fontsize=14)
        ax.set_ylabel("Lift Counts", fontsize=14)
        ax.set_title(r"Training Points vs. Lift Count for $C_L$", fontsize=15)

        plt.savefig(f'Convergence_Cl')
        plt.close()

        _, ax = plt.subplots()
        ax.plot(self.train_pts_list, self.cd_count_list, marker="o")
        ax.set_xlabel("Training Points", fontsize=14)
        ax.set_ylabel("Drag Count", fontsize=14)
        ax.set_title(r"Training Points vs. Drag Count for $C_D$", fontsize=15)

        plt.savefig(f'Convergence_Cd')
        plt.close()

        _, ax = plt.subplots()
        ax.plot(self.train_pts_list, self.qdot_rmse_list, marker="o")
        ax.set_xlabel("Training Points", fontsize=14)
        ax.set_ylabel("QDot NRMSE", fontsize=14)
        ax.set_title(r"Training Points vs. $\dot{Q}$ NRMSE", fontsize=15)

        plt.savefig(f'Convergence_QDot')
        plt.close()

        logging.info("Finished creating convergence plots")

    def write_convergence_data(self):
        # Create text file to store info if user wants to compare or see what's going on during convergence
        table_data = [["Num Training Points", "Lift Count", "Drag Count", "QDot NRMSE"]]

        # Create empty list, append training pt number, cl count number, cd count number to that list then append
        # entire list to table data, i.e. this temp list will be a row of output
        for i in range(len(self.train_pts_list)):
            temp = []
            temp.append(self.train_pts_list[i])
            temp.append(self.cl_count_list[i])
            temp.append(self.cd_count_list[i])
            temp.append(self.qdot_rmse_list[i])
            table_data.append(temp)

        # Create & write to data file
        with open('num_train_pts_vs_counts.txt', 'w') as file:

            for row in table_data:
                file.write("{:^20} {:^20} {:^20} {:^20}".format(*row))
                file.write('\n')

        logging.info("Finished creating convergence data file")

    def write_point_data(self, x, cl, cd, qdot, num_training_pts, qdot_opt):
        if qdot_opt == QDOT_CB2OTIS:
            table_data_train = [["Alpha (°)", "Mach Number", "Cl", "Cd"]]
            table_data_test = [["Alpha (°)", "Mach Number", "Cl", "Cd"]]

            for i in range(len(x.train)):
                train_temp = []
                train_temp.append(x.train[i][0])
                train_temp.append(x.train[i][1])
                train_temp.append(cl.train[i][0])
                train_temp.append(cd.train[i][0])
                table_data_train.append(train_temp)

            # Create and write to the data file
            with open('train_pts_data_cl&cd_' + str(num_training_pts) + 'pts.txt', 'w') as file:

                for row in table_data_train:
                    file.write("{:^20} {:^20} {:^20} {:^20}".format(*row))
                    file.write('\n')

            for i in range(len(x.test)):
                test_temp = []
                test_temp.append(x.test[i][0])
                test_temp.append(x.test[i][1])
                test_temp.append(cl.test[i][0])
                test_temp.append(cd.test[i][0])
                table_data_test.append(test_temp)

            # Create and write to the data file
            with open('test_pts_data_cl&cd_' + str(num_training_pts) + 'pts.txt', 'w') as file:

                for row in table_data_test:
                    file.write("{:^20} {:^20} {:^20} {:^20}".format(*row))
                    file.write('\n')

            logging.info("Finished creating training & testing pts data file")

        elif qdot_opt == QDOT_ADB2TECPLOT:
            table_data_train = [["Alpha (°)", "Mach Number", "Dynamic Pressure (bars)", "Qdot"]]
            table_data_test = [["Alpha (°)", "Mach Number", "Dynamic Pressure (bars)", "Qdot"]]

            for i in range(len(x.train)):
                train_temp = []
                train_temp.append(x.train[i][0])
                train_temp.append(x.train[i][1])
                train_temp.append(x.train[i][2])
                train_temp.append(qdot.train[i][0])
                table_data_train.append(train_temp)

            # Create and write to the data file
            with open('train_pts_data_qdot_' + str(num_training_pts) + 'pts.txt', 'w') as file:

                for row in table_data_train:
                    file.write("{:^20} {:^20} {:^20} {:^20}".format(*row))
                    file.write('\n')

            for i in range(len(x.test)):
                test_temp = []
                test_temp.append(x.test[i][0])
                test_temp.append(x.test[i][1])
                test_temp.append(x.test[i][2])
                test_temp.append(qdot.test[i][0])
                table_data_test.append(test_temp)

            # Create and write to the data file
            with open('test_pts_data_qdot_' + str(num_training_pts) + 'pts.txt', 'w') as file:

                for row in table_data_test:
                    file.write("{:^20} {:^20} {:^20} {:^20}".format(*row))
                    file.write('\n')

            logging.info("Finished creating training & testing pts data file")


    def plot_lhs_training(self, x_values_train, num_training_pts, qdot_opt):
        alpha_train_lhs = []
        mach_train_lhs = []
        q_train_lhs = []

        # Get list for all alphas, mach numbers, and q's from LHS pts (i.e. our training data set)
        for i in range(len(x_values_train)):
            alpha_train_lhs.append(x_values_train[i][0])
            mach_train_lhs.append(x_values_train[i][1])

            if qdot_opt == QDOT_ADB2TECPLOT:
                q_train_lhs.append(x_values_train[i][2])

        if qdot_opt == QDOT_ADB2TECPLOT:

            # Plot LHS scheme on 3D plot to show coverage of design space
            ax = plt.axes(projection='3d')
            ax.scatter3D(alpha_train_lhs, mach_train_lhs, q_train_lhs, 'blue')
            ax.set_xlabel('Angle of Attack (degrees)')
            ax.set_ylabel('Mach Number')
            ax.set_zlabel('Dynamic Pressure (bars)')
            ax.set_title('LHS for Training Points')
            ax.set_xlim(self.min_alpha, self.max_alpha)
            ax.set_ylim(self.min_mach, self.max_mach)
            ax.set_zlim(self.min_q, self.max_q)
            ax.view_init(10, -45)

            plt.savefig(f'LHS_TrainingPTS_Qdot_{num_training_pts}pts')
            plt.close()

        else:
            # Plot LHS scheme on @D plot to show coverage of design space
            _, ax = plt.subplots()
            ax.scatter(alpha_train_lhs, mach_train_lhs)
            ax.set_xlabel('Angle of Attack (degrees)')
            ax.set_ylabel('Mach Number')
            ax.set_title('LHS for Training Points')
            ax.set_xlim(self.min_alpha, self.max_alpha)
            ax.set_ylim(self.min_mach, self.max_mach)

            plt.savefig(f'LHS_TrainingPTS_Cl&Cd_{num_training_pts}pts')
            plt.close()

        logging.info("Finished creating LHS plot for training data")

    def plot_lhs_training_cokrig(self, x_values_train, x_hf, num_training_pts):
        alpha_train_lhs = []
        mach_train_lhs = []
        a_hf = []
        m_hf = []

        # Get list for all alphas, mach numbers, and q's from LHS pts (i.e. our training data set)

        for i in range(len(x_values_train)):
            alpha_train_lhs.append(x_values_train[i][0])
            mach_train_lhs.append(x_values_train[i][1])

        for i in range(len(x_hf)):
            a_hf.append(x_hf[i][0])
            m_hf.append(x_hf[i][1])

        # Plot LHS scheme on @D plot to show coverage of design space
        _, ax = plt.subplots()
        ax.scatter(alpha_train_lhs, mach_train_lhs, label='LF Points')
        ax.scatter(a_hf, m_hf, color='orange', label='HF Points')
        ax.set_xlabel('Angle of Attack (degrees)')
        ax.set_ylabel('Mach Number')
        ax.set_title('LHS for Training Points')
        ax.set_xlim(self.min_alpha, self.max_alpha)
        ax.set_ylim(self.min_mach, self.max_mach)
        ax.legend()

        plt.savefig(f'LHS_TrainingPTS_Cokrig_Cl&Cd_{num_training_pts}pts')
        plt.close()

        logging.info("Finished creating LHS plot for training data")

    def copy_model(self):
        """Copy relevant model files and dir into otis dir"""

        for filename in [self.model_filename, self.model_name+MESH_EXT, self.model_name+SETUP_EXT]:
            shutil.copy(os.path.join(self.model_dir, filename),
                        os.path.join(self.otis_dir, filename))

        shutil.copytree(os.path.join(self.model_dir, TAG_REGIONS),
                        os.path.join(self.otis_dir, TAG_REGIONS))

    def clear_lists(self):
        self.train_pts_list = []
        self.cl_count_list = []
        self.cd_count_list = []
        self.qdot_rmse_list = []

    def update_lists(self, num_training_pts, lift_count, drag_count, qdot_count):
        """Save lift, drag & qdot counts to file in otis dir."""
        self.append_lists(num_training_pts, lift_count, drag_count, qdot_count)
        path = os.path.join(self.otis_dir, PTS_CL_CD_FILE)

        with open(path, "wb") as f:
            pickle.dump((num_training_pts, lift_count, drag_count, qdot_count), f)

        logging.info(f'Recorded cl/cd/qdot counts to file: {path}')

    def append_lists(self, num_training_pts, lift_count, drag_count, qdot_count):
        """Add an entry to each list."""
        self.train_pts_list.append(num_training_pts)
        self.cl_count_list.append(lift_count)
        self.cd_count_list.append(drag_count)
        self.qdot_rmse_list.append(qdot_count)

    def get_list_entry(self):
        """Read lift and drag counts from saved data files."""
        with open(os.path.join(self.otis_dir, PTS_CL_CD_FILE), "rb") as f:
            return pickle.load(f)

    def run(self, start, stop, step, run_type=SERIAL, task_id=0):
        """Execute CBAero-surrogate processing based on run type."""

        x = TrainTestValues()
        x_lhs = TrainTestValues()
        x_lhs_qdot = TrainTestValues()
        cl = TrainTestValues()
        cd = TrainTestValues()
        qdot = TrainTestValues()
        hf = None

        if run_type in [SERIAL, PARALLEL_END]:
            self.copy_model()

        if self.cokrig_file is not None:
            # Read data out of cokrig file
            hf_data = cbas.get_hf_data()

            # Set test values
            x.test = hf_data.x.test
            cd.test = hf_data.cd.test
            cl.test = hf_data.cl.test

            # Set hf train values
            hf = HighFidelityTrain(x=hf_data.x.train, cd=hf_data.cd.train, cl=hf_data.cl.train)

        # Loop on num train pt values, use index of value to match up w/task when running in parallel
        for i, pts in enumerate(range(start, stop, step)):

            # Check if we should continue with this loop iteration
            if run_type == PARALLEL_START:

                if i == int(task_id):  # NOTE Assumes job array starts at zero!
                    # Parallel only: use otis subdir based on task ID - NOTE: non-parallel: use base otis dir
                    self.change_otis_dir(os.path.join(self.base_otis_dir, str(i)))
                    self.copy_model()
                else:
                    continue   # Skip this loop iteration b/c this task is meant for another train pts value

            elif run_type == PARALLEL_END:  # Doing check here (inside loop) to avoid more indenting
                continue
            else:  # run_type is SERIAL, so we need to run the code below
                pass

            # Create latin hypercube, run CBAero. Running in 1 of 2 modes: 1) cokrig for cl/cd and krig for heat flux, or 2) krig for all three

            # If only the cl and cd come from cokrig_file, but not qdot. Need to run separate cases for qdot.
            if self.cokrig_file is not None:
                # Cl and Cd
                x_lhs.train = self.get_latin_hypercube_cokriging(pts, RAND_STATE)
                self.run_cbaero(TRAIN, pts, x_lhs.train, QDOT_CB2OTIS, task_id)

                # Qdot
                x_lhs_qdot.train, x_lhs_qdot.test = self.latin_hypercube_qdot(pts, RAND_STATE)
                self.run_cbaero(TRAIN, pts, x_lhs_qdot.train, QDOT_ADB2TECPLOT, task_id)
                self.run_cbaero(TEST, int(pts * TEST_TO_TRAIN), x_lhs_qdot.test, QDOT_ADB2TECPLOT, task_id)

                x.train, cl.train, cd.train = self.extract_otis_cokrig(pts)
                sm_cl, lift_count = self.fit_cokrig(CL_MODE, x.train, hf.x, x.test, cl.train, hf.cl, cl.test, pts)
                sm_cd, drag_count = self.fit_cokrig(CD_MODE, x.train, hf.x, x.test, cd.train, hf.cd, cd.test, pts)

                _, _, _, qdot = self.extract_otis(pts, QDOT_CB2OTIS)
                sm_qdot, qdot_count = self.fit_krig(QDOT_MODE, x_lhs_qdot, qdot, CORR_FUNC, pts, QDOT_CB2OTIS)

                self.write_point_data(x, cl, cd, qdot, pts, QDOT_CB2OTIS)
                self.plot_lhs_training_cokrig(x.train, hf.x, pts)

            # If there is no cokrig_file, all surrogates will be kriging
            else:
                # Cl and Cd
                x_lhs.train, x_test = self.latin_hypercube(pts, RAND_STATE)
                self.run_cbaero(TRAIN, pts, x_lhs.train, QDOT_CB2OTIS, task_id)
                self.run_cbaero(TEST, int(pts * TEST_TO_TRAIN), x_test, QDOT_CB2OTIS, task_id)

                # Qdot
                x_lhs_qdot.train, x_lhs_qdot.test = self.latin_hypercube_qdot(pts, RAND_STATE)
                self.run_cbaero(TRAIN, pts, x_lhs_qdot.train, QDOT_ADB2TECPLOT, task_id)
                self.run_cbaero(TEST, int(pts * TEST_TO_TRAIN), x_lhs_qdot.test, QDOT_ADB2TECPLOT, task_id)

                x, cl, cd, qdot = self.extract_otis(pts, QDOT_BOTH)
                sm_cl, lift_count = self.fit_krig(CL_MODE, x, cl, CORR_FUNC, pts, QDOT_BOTH)
                sm_cd, drag_count = self.fit_krig(CD_MODE, x, cd, CORR_FUNC, pts, QDOT_BOTH)
                sm_qdot, qdot_count = self.fit_krig(QDOT_MODE, x_lhs_qdot, qdot, CORR_FUNC, pts, QDOT_CB2OTIS)
                self.write_point_data(x, cl, cd, qdot, pts, QDOT_CB2OTIS)
                self.plot_lhs_training(x.train, pts, QDOT_CB2OTIS)

            self.write_point_data(x_lhs_qdot, cl, cd, qdot, pts, QDOT_ADB2TECPLOT)
            self.plot_lhs_training(x_lhs_qdot.train, pts, QDOT_ADB2TECPLOT)

            logging.info(f'sm_cl: {sm_cl}, sm_cd {sm_cd}, sm_qdot: {sm_qdot}')
            self.update_lists(pts, lift_count, drag_count, qdot_count)  # Writes to file & maintain lists

        # Loop has ended, do postprocessing...

        if run_type in [SERIAL, PARALLEL_END]:

            if run_type == PARALLEL_END:
                # Reassemble lists of training points, cl/cd/qdots counts
                self.clear_lists()

                for i, pts in enumerate(range(start, stop, step)):
                    self.change_otis_dir(os.path.join(self.base_otis_dir, str(i)))
                    pts, lift_count, drag_count, qdot_count = self.get_list_entry()
                    self.append_lists(pts, lift_count, drag_count, qdot_count)

                self.change_otis_dir(BASE_OTIS_DIR)

            logging.info(f'Training Point List: {self.train_pts_list}')
            logging.info(f'Lift Count List: {self.cl_count_list}')
            logging.info(f'Drag Count List: {self.cd_count_list}')
            logging.info(f'QDot RMSE List: {self.qdot_rmse_list}')

            self.plot_convergence()  # Plot convergence for lift & drag counts
            self.write_convergence_data()  # Save conv & pts to files

if __name__ == "__main__":

    # Command line arugments

    parser = argparse.ArgumentParser(description='Get the arguments from command line for Surrogate Model Script')

    # ...Model file - include path & ext
    parser.add_argument('model_file', type=str,
                        help='The path, name, and extension of the model file.')
    # ...Training points range
    parser.add_argument('start', type=int,
                        help='The number of training points to start with.')
    parser.add_argument('stop', type=int,
                        help='The number of training points to finish with if you are testing for convergence and want to test many \
                              different amounts of training points to see how many is good enough.')
    parser.add_argument('step', type=int,
                        default=10,
                        help='If you are testing different amounts of training points to see how many is good enough, \
                              this will be the step size of each iteration. If your starting number of training points is 25, your final number 126, and your step \
                              size is 25, the code will create a surrogate model using 25 training points, then 50, then 75, 100, and finally 125 and plot the error \
                              associated with each. The user can then see how many training points to use to create a model that is "good enough". If you are not testing \
                              multiple sets, simply set the step size to be greater than or equal to the interval between "start" and "stop." For example, setting \
                              the start to be 400, stop to be 401, and step to be 1 will mean the code will only run once with 400 training points.')
    # Run type
    parser.add_argument('--run_type', type=str,
                        default=SERIAL, choices=[SERIAL, PARALLEL_START, PARALLEL_END],
                        help='The type of run being used. There are three types: \
                              "serial", "pstart", "pend". "serial" will generate Latin Hypercube samples, run CBAero for each point, and save the CBAero output as .otis files \
                              before reading those files and building a Kriging Surrogate. \
                              If the user provides a data file (high fidelity data), the code will build cokriging surrogate models using the given data as "high fidelity" and \
                              CBAero as "low fidelity". "serial" will build kriging or cokriging models depending on if you provide a data file or not. \
                              "pstart" runs "serial" logic, in parallel, only for the training points indexed by "task_id". \
                              "pend" runs the model building processes after CBAero has been run (logic on data produced using "pstart"). \
                              "serial" is the default value.')
    # ...Parameter values
    parser.add_argument('--min_alpha', type=float, default=0,
                        help='The minumum angle of attack (in degrees) to test over. Default = 0')
    parser.add_argument('--max_alpha', type=float, default=30,
                        help='The maximum angle of attack (in degrees) to test over. Default = 30')
    parser.add_argument('--min_mach', type=float, default=1.2,
                        help='The minimum mach number to test over. It cannot be less than 1.2. Default = 1.2')
    parser.add_argument('--max_mach', type=float, default=25,
                        help='The maximum mach number to test over. Default = 25')
    parser.add_argument('--min_q', type=float, default=0.0035,
                        help='The minimum dynamic pressure (in bars) to test over. Default = 0.0035')
    parser.add_argument('--max_q', type=float, default=10,
                        help='The maximum dynamic pressure (in bars) to test over. Default = 10')
    #...Running CoKriging?
    parser.add_argument('--cokrig_file', type=str,
                        default=None,
                        help='The path to the high fidelity training data file.')
    # ...Save Krigging model?
    parser.add_argument('--save_as', type=str,
                        default=None,
                        help='This will be used to form the name of the save file for the Kriging Surrogate Models. There are two surrogates \
                              that will be saved: one for the lift coefficient (Cl) and one for the drag coefficient (Cd). "_cl" and "_cd" will be appended to the filename you input \
                              to differentiate the two models. For instance, if you use the default file name, you will have two files in the current directory: "kriging_cl.pkl" and \
                              kriging_cd.pkl". If you use your own name (maybe to distinguish the model geometry you are using), it will be "[user_input]_cl.pkl" for the Cl model. \
                              The number of training points will also be appended to the file name, thus saving each model during a convergence study. "[user_input]_cl_[num_training_pts]pts.pkl" \
                                Note that these files will be pkl files (we are using pickle for the save feature). ')
    # ...Parallel?
    parser.add_argument('--task_id', type=int,
                        required='--run_type=start' in str(sys.argv),
                        default=None,
                        help='Current task ID if running job array (usually $SLURM_ARRAY_TASK_ID). \
                              Only valid when "run_type" is "start".')

    args = parser.parse_args()

    # Ensure task ID provided when running in parallel
    if args.run_type == PARALLEL_START and args.task_id is None:
        raise Exception('error: argument "--run_type=pstart" requires argument "--task_id=<task_number>"!')

    # Set up logging
    logging.basicConfig(format=f'%(levelname)s: %(message)s (%(filename)s:%(lineno)d, %(asctime)s, task:{args.task_id})',
                        stream=sys.stdout,
                        level=logging.DEBUG),
    logging.getLogger('matplotlib.font_manager').setLevel(logging.WARNING)  # Suppress debug entries for "...findfont..."
    logging.info('Start')

    # Create object to manage CBAero surrogate modeling
    cbas = CbaeroSurrogate(args.model_file,
                           args.min_alpha, args.max_alpha,
                           args.min_mach, args.max_mach,
                           args.min_q, args.max_q,
                           args.cokrig_file, args.save_as)

    # Build surrogate model
    cbas.run(args.start, args.stop, args.step, args.run_type, args.task_id)  

    logging.info(f'End')
